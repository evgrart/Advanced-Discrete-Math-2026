from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / ".deps"))

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from build_gradebook import extract_roster_and_sample


ROOT = Path(__file__).parent
OUTPUT = ROOT / "ДМ_система_логи.xlsx"
SOURCE = next(p for p in ROOT.glob("*.xlsx") if p.name == "ДМ Артём.xlsx")

PRACTICE_COUNT = 15
STUDENTS_PER_GROUP = 30
TASK_COUNT = 30
PLUS_BLOCK_SIZE = 68

NAVY = "17365D"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
RED = "C00000"
LIGHT_RED = "FCE4D6"
WHITE = "FFFFFF"
GRAY = "D9E1F2"
INPUT = "FFFFFF"
THIN = Side(style="thin", color="D9E1F2")


def title(ws, text, subtitle, end_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = text
    ws["A1"].font = Font(size=18, bold=True, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30


def header_style(ws, row, end_col):
    for col in range(1, end_col + 1):
        c = ws.cell(row, col)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color=WHITE))
    ws.row_dimensions[row].height = 28


def body_style(ws, first_row, last_row, first_col, last_col):
    for row in ws.iter_rows(min_row=first_row, max_row=last_row, min_col=first_col, max_col=last_col):
        for c in row:
            c.border = Border(bottom=THIN)
            c.alignment = Alignment(vertical="center", wrap_text=True)


def color_plus_sheet(ws, header_row, first_row, last_row, task_start, task_end):
    # Screenshot-like header colors: name, TG, total, tasks.
    for col, fill in [(1, "A9D18E"), (2, "9DC3E6"), (3, "F4B183")]:
        ws.cell(header_row, col).fill = PatternFill("solid", fgColor=fill)
        ws.cell(header_row, col).font = Font(bold=True, color="000000")
    for col in range(task_start, task_end + 1):
        ws.cell(header_row, col).fill = PatternFill("solid", fgColor="38761D")
        ws.cell(header_row, col).font = Font(bold=True, color=WHITE)

    rng = f"{get_column_letter(task_start)}{first_row}:{get_column_letter(task_end)}{last_row}"
    anchor = f"{get_column_letter(task_start)}{first_row}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{anchor}="+"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{anchor}="1"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{anchor}="-"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=WHITE)))


def add_validation(ws, cell_range, values, message):
    validation = DataValidation(type="list", formula1='"' + values + '"', allow_blank=True)
    validation.errorStyle = "stop"
    validation.showErrorMessage = True
    validation.errorTitle = "Недопустимое значение"
    validation.error = message
    ws.add_data_validation(validation)
    validation.add(cell_range)


def build_cw_sheet(ws, source_path):
    source_wb = load_workbook(source_path, data_only=False)
    source = source_wb.worksheets[0]
    title(ws, "CW — контрольные работы", "Структура перенесена из листа «КР»: после имени добавлен столбец «Попытка».", 11)
    ws.unmerge_cells("A1:K1")
    ws.unmerge_cells("A2:K2")
    ws["A2"] = None
    ws["K1"] = "CW"
    ws["K1"].font = Font(bold=True, color=NAVY)
    ws["K2"] = "Попытка — номер текущей сдачи"
    ws["K2"].font = Font(italic=True, color="666666")

    section_starts = []
    for row in range(1, source.max_row + 1):
        if source.cell(row, 4).value and source.cell(row, 1).value is None and source.cell(row, 3).value is None:
            section_starts.append(row)

    for index, start in enumerate(section_starts):
        next_start = section_starts[index + 1] if index + 1 < len(section_starts) else source.max_row + 1
        title_row = start
        weight_row = start + 1
        header_row = start + 2
        first_student = start + 3
        last_student = next_start - 1

        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=9)
        ws.cell(title_row, 1, source.cell(title_row, 4).value)
        ws.cell(title_row, 1).font = Font(size=14, bold=True, color=NAVY)
        ws.cell(title_row, 1).fill = PatternFill("solid", fgColor="D9D9D9")
        ws.cell(title_row, 1).alignment = Alignment(horizontal="center")

        ws.cell(weight_row, 4, "W")
        for col in range(4, 9):
            ws.cell(weight_row, col + 1, source.cell(weight_row, col).value)
        ws.cell(header_row, 1, "Имя")
        ws.cell(header_row, 2, "Попытка")
        ws.cell(header_row, 3, "TG")
        ws.cell(header_row, 4, "S")
        for col in range(4, 9):
            ws.cell(header_row, col + 1, source.cell(header_row, col).value)
        header_style(ws, header_row, 11)
        for col, fill in [(1, "A9D18E"), (2, "FFF2CC"), (3, "9DC3E6"), (4, "F4B183")]:
            ws.cell(header_row, col).fill = PatternFill("solid", fgColor=fill)
            ws.cell(header_row, col).font = Font(bold=True, color="000000")
        for col in range(5, 10):
            ws.cell(header_row, col).fill = PatternFill("solid", fgColor="38761D")
            ws.cell(header_row, col).font = Font(bold=True, color=WHITE)

        for row in range(first_student, last_student + 1):
            if source.cell(row, 1).value:
                ws.cell(row, 1, source.cell(row, 1).value)
                ws.cell(row, 2).fill = PatternFill("solid", fgColor="FFF2CC")
                ws.cell(row, 2).protection = Protection(locked=False)
                ws.cell(row, 3, source.cell(row, 2).value)
                ws.cell(row, 4, f'=IF(A{row}="","",SUMPRODUCT(E{row}:I{row},E${weight_row}:I${weight_row}))')
                for col in range(5, 10):
                    ws.cell(row, col, None)
                    ws.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
                    ws.cell(row, col).protection = Protection(locked=False)
                ws.cell(row, 10, None)
                ws.cell(row, 11, None)
            else:
                for col in range(1, 12):
                    if source.cell(row, col if col <= 3 else col - 1).value is not None:
                        ws.cell(row, col, source.cell(row, col if col <= 3 else col - 1).value)
        body_style(ws, first_student, last_student, 1, 11)
        add_validation(ws, f"B{first_student}:B{last_student}", "1,2,3,4", "Укажите номер попытки от 1 до 4")
        score_validation = DataValidation(type="decimal", operator="between", formula1="0", formula2="4", allow_blank=True)
        score_validation.errorStyle = "stop"
        score_validation.showErrorMessage = True
        score_validation.errorTitle = "Недопустимый балл"
        score_validation.error = "Введите число от 0 до 4"
        ws.add_data_validation(score_validation)
        score_validation.add(f"E{first_student}:I{last_student}")

    ws.freeze_panes = "E4"
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = False
    for col, width in {"A": 27, "B": 12, "C": 22, "D": 12, "J": 22, "K": 22}.items():
        ws.column_dimensions[col].width = width
    for col in range(5, 10):
        ws.column_dimensions[get_column_letter(col)].width = 11


def build_cw_sheet_template(ws, source_path, students):
    source_wb = load_workbook(source_path, data_only=False)
    source = source_wb.worksheets[0]
    ws["K1"] = "CW"
    ws["K1"].font = Font(bold=True, color=NAVY)
    ws["K2"] = "Попытка — номер текущей сдачи"
    ws["K2"].font = Font(italic=True, color="666666")

    section_starts = []
    for row in range(1, source.max_row + 1):
        if source.cell(row, 4).value and source.cell(row, 1).value is None and source.cell(row, 3).value is None:
            section_starts.append(row)

    output_row = 1
    for start in section_starts:
        title_row, weight_row, header_row = output_row, output_row + 1, output_row + 2
        first_student = output_row + 3
        last_student = first_student + len(students) - 1
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=9)
        ws.cell(title_row, 1, source.cell(start, 4).value)
        ws.cell(title_row, 1).font = Font(size=14, bold=True, color=NAVY)
        ws.cell(title_row, 1).fill = PatternFill("solid", fgColor="D9D9D9")
        ws.cell(title_row, 1).alignment = Alignment(horizontal="center")
        ws.cell(weight_row, 4, "W")
        for col in range(4, 9):
            ws.cell(weight_row, col + 1, source.cell(start + 1, col).value)
        headers = ["Имя", "Попытка", "TG", "S"] + [source.cell(start + 2, col).value for col in range(4, 9)]
        for col, value in enumerate(headers, start=1):
            ws.cell(header_row, col, value)
        header_style(ws, header_row, 11)
        for col, fill in [(1, "A9D18E"), (2, "FFF2CC"), (3, "9DC3E6"), (4, "F4B183")]:
            ws.cell(header_row, col).fill = PatternFill("solid", fgColor=fill)
            ws.cell(header_row, col).font = Font(bold=True, color="000000")
        for col in range(5, 10):
            ws.cell(header_row, col).fill = PatternFill("solid", fgColor="38761D")
            ws.cell(header_row, col).font = Font(bold=True, color=WHITE)
        for row, (name, tg) in enumerate(students, start=first_student):
            ws.cell(row, 1, name)
            ws.cell(row, 2).fill = PatternFill("solid", fgColor="FFF2CC")
            ws.cell(row, 2).protection = Protection(locked=False)
            ws.cell(row, 3, tg)
            ws.cell(row, 4, f'=IF(A{row}="","",SUMPRODUCT(E{row}:I{row},E${weight_row}:I${weight_row}))')
            for col in range(5, 10):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
                ws.cell(row, col).protection = Protection(locked=False)
        body_style(ws, first_student, last_student, 1, 11)
        add_validation(ws, f"B{first_student}:B{last_student}", "1,2,3,4", "Укажите номер попытки от 1 до 4")
        score_validation = DataValidation(type="decimal", operator="between", formula1="0", formula2="4", allow_blank=True)
        score_validation.errorStyle = "stop"
        score_validation.showErrorMessage = True
        score_validation.errorTitle = "Недопустимый балл"
        score_validation.error = "Введите число от 0 до 4"
        ws.add_data_validation(score_validation)
        score_validation.add(f"E{first_student}:I{last_student}")
        output_row = last_student + 2

    ws.freeze_panes = "E4"
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 27, "B": 12, "C": 22, "D": 12, "J": 22, "K": 22}.items():
        ws.column_dimensions[col].width = width
    for col in range(5, 10):
        ws.column_dimensions[get_column_letter(col)].width = 11


def build_ladder_sheet(ws):
    title(ws, "Логи D", "Служебный журнал выходов. Записи добавляются автоматически, когда практик заменяет + на 1 в верхнем плюсовике.", 10)
    ws["A3"] = "Практик (1 Артём / 2 Рами / 3 Немат)"
    ws["B3"] = 1
    ws["C3"] = "Студент"
    ws["D3"] = ""
    ws["E3"] = "Задача"
    ws["F3"] = ""
    ws["G3"] = "Практика"
    ws["H3"] = 2
    ws["I3"] = "Разрешить выход после 15"
    ws["J3"] = "Нет"
    for col in [2, 4, 6, 8, 10]:
        ws.cell(3, col).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(3, col).protection = Protection(locked=False)
    ws["A4"] = "Кнопка учитывает только задачи, где у студента стоит + в таблице выбранной практики, и не выдаёт больше двух выходов за одну практику. После 15 выходов требуется явное разрешение."
    ws.merge_cells("A4:J4")
    ws["A4"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[4].height = 32
    headers = ["Практика", "Практикант", "Студент", "Задача", "Номер выхода", "База D", "Коэффициент", "Начислено D", "Статус", "Поправка"]
    for col, value in enumerate(headers, start=1):
        ws.cell(6, col, value)
    header_style(ws, 6, 10)
    for row in range(7, 3007):
        ws.cell(row, 5, f'=IF(C{row}="","",COUNTIF($C$7:C{row},C{row}))')
        ws.cell(row, 6, f'=IF(E{row}="","",IF(E{row}<=3,6,IF(E{row}<=6,5,IF(E{row}<=9,4,IF(E{row}<=12,3,IF(E{row}<=15,2,0))))))')
        ws.cell(row, 8, f'=IF(OR(F{row}="",G{row}=""),"",F{row}*G{row}+IF(J{row}="",0,J{row}))')
        for col in [1, 2, 3, 4, 7, 9, 10]:
            ws.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
            ws.cell(row, col).protection = Protection(locked=False)
    body_style(ws, 7, 3006, 1, 10)
    add_validation(ws, "B3", "1,2,3", "Выберите практиканта 1, 2 или 3")
    add_validation(ws, "J3", "Нет,Да", "Выберите Нет или Да")
    for col in ["H3"]:
        ws[col].data_type = "n"
    ws.freeze_panes = "A7"
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 12, "B": 15, "C": 27, "D": 16, "E": 14, "F": 10, "G": 14, "H": 14, "I": 24, "J": 12}.items():
        ws.column_dimensions[col].width = width
    for col in ["L", "M", "N", "O"]:
        ws.column_dimensions[col].hidden = True
    ws.row_dimensions[3].hidden = True
    ws.row_dimensions[4].hidden = True


def build_plus_sheet(ws, practitioner, roster, task_labels, task_values, result_values, common_config_ref, practitioner_no):
    """Create 15 identical practice blocks on one practitioner sheet."""
    end_task_col = 35  # F:AI = 30 tasks
    task_labels = [str(x) for x in task_labels[:TASK_COUNT]]
    task_labels += [str(i) for i in range(len(task_labels) + 1, TASK_COUNT + 1)]
    template_roster = roster[:STUDENTS_PER_GROUP] + [("", "")] * max(0, STUDENTS_PER_GROUP - len(roster))

    title(ws, f"Плюсы — {practitioner}", "В каждом блоке одна практика. В верхней таблице ставятся +; практик может заменить + на 1. В нижней таблице практик отмечает 1 или -.", end_task_col)

    for practice_no in range(1, PRACTICE_COUNT + 1):
        block_start = 4 + (practice_no - 1) * PLUS_BLOCK_SIZE
        plus_header = block_start + 1
        plus_first = block_start + 2
        plus_last = plus_first + STUDENTS_PER_GROUP - 1
        result_title = block_start + 33
        result_header = block_start + 34
        result_first = result_header + 1
        result_last = result_first + STUDENTS_PER_GROUP - 1

        ws.merge_cells(start_row=block_start, start_column=1, end_row=block_start, end_column=end_task_col)
        ws.cell(block_start, 1, f"Практика {practice_no} — плюсы")
        ws.cell(block_start, 1).font = Font(size=14, bold=True, color=NAVY)
        ws.cell(block_start, 1).fill = PatternFill("solid", fgColor="D9D9D9")

        headers = ["Имя", "TG", "S", "Коэффициент", "D за практику"] + task_labels
        for col, value in enumerate(headers, start=1):
            ws.cell(plus_header, col, value)
        header_style(ws, plus_header, end_task_col)
        for col, fill in [(1, "A9D18E"), (2, "9DC3E6"), (3, "F4B183"), (4, "FFF2CC"), (5, "D9EAD3")]:
            ws.cell(plus_header, col).fill = PatternFill("solid", fgColor=fill)
            ws.cell(plus_header, col).font = Font(bold=True, color="000000")
        for col in range(6, end_task_col + 1):
            ws.cell(plus_header, col).fill = PatternFill("solid", fgColor="38761D")
            ws.cell(plus_header, col).font = Font(bold=True, color=WHITE)

        for row, (name, tg) in enumerate(template_roster, start=plus_first):
            ws.cell(row, 1, name)
            ws.cell(row, 2, tg)
            ws.cell(row, 3, f'=IF(A{row}="","",COUNTIF(F{row}:AI{row},"+")+COUNTIF(F{row}:AI{row},"1"))')
            ws.cell(row, 4, f'=IF(A{row}="","",IF(C{row}>2*COUNTA(F${plus_header}:AI${plus_header})/3,\'Общий\'!$M$11,IF(C{row}>=COUNTA(F${plus_header}:AI${plus_header})/2,\'Общий\'!$M$12,\'Общий\'!$M$13)))')
            ws.cell(row, 5, f'=IF(A{row}="","",SUMIFS(\'Логи\'!$H$7:$H$3006,\'Логи\'!$A$7:$A$3006,{practice_no},\'Логи\'!$B$7:$B$3006,{practitioner_no},\'Логи\'!$C$7:$C$3006,A{row}))')
            for col in range(6, end_task_col + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
                ws.cell(row, col).protection = Protection(locked=False)
        body_style(ws, plus_first, plus_last, 1, end_task_col)
        add_validation(ws, f"F{plus_first}:AI{plus_last}", "+,1", "Студент ставит +; практик может поставить 1")
        ws.conditional_formatting.add(f"C{plus_first}:C{plus_last}", CellIsRule(operator="greaterThanOrEqual", formula=["21"], fill=PatternFill("solid", fgColor=GREEN)))
        ws.conditional_formatting.add(f"C{plus_first}:C{plus_last}", CellIsRule(operator="between", formula=["15", "20"], fill=PatternFill("solid", fgColor="A9D18E")))
        ws.conditional_formatting.add(f"C{plus_first}:C{plus_last}", CellIsRule(operator="lessThan", formula=["15"], fill=PatternFill("solid", fgColor=LIGHT_GREEN)))
        color_plus_sheet(ws, plus_header, plus_first, plus_last, 6, end_task_col)

        ws.merge_cells(start_row=result_title, start_column=1, end_row=result_title, end_column=end_task_col)
        ws.cell(result_title, 1, f"Практика {practice_no} — результаты рассказа")
        ws.cell(result_title, 1).font = Font(size=12, bold=True, color=NAVY)
        ws.cell(result_title, 1).fill = PatternFill("solid", fgColor="D9EAD3")
        result_headers = ["Имя", "TG", "", "", ""] + task_labels
        for col, value in enumerate(result_headers, start=1):
            ws.cell(result_header, col, value)
        header_style(ws, result_header, end_task_col)
        for col, fill in [(1, "A9D18E"), (2, "9DC3E6")]:
            ws.cell(result_header, col).fill = PatternFill("solid", fgColor=fill)
            ws.cell(result_header, col).font = Font(bold=True, color="000000")
        for col in range(6, end_task_col + 1):
            ws.cell(result_header, col).fill = PatternFill("solid", fgColor="C00000")
            ws.cell(result_header, col).font = Font(bold=True, color=WHITE)
        for row, (name, tg) in enumerate(template_roster, start=result_first):
            ws.cell(row, 1, name)
            ws.cell(row, 2, tg)
            for col in range(6, end_task_col + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
                ws.cell(row, col).protection = Protection(locked=False)
        body_style(ws, result_first, result_last, 1, end_task_col)
        add_validation(ws, f"F{result_first}:AI{result_last}", "1,-", "Практик может ввести только 1, - или оставить пусто")
        color_plus_sheet(ws, result_header, result_first, result_last, 6, end_task_col)

    ws.protection.sheet = False
    ws.freeze_panes = "F6"
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 27, "B": 22, "C": 9, "D": 14, "E": 14}.items():
        ws.column_dimensions[col].width = width
    for col in range(6, end_task_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10


def build():
    artem = "\u0410\u0440\u0442\u0451\u043c"
    rami = "\u0420\u0430\u043c\u0438"
    nemat = "\u041d\u0435\u043c\u0430\u0442"
    group_rosters = {
        artem: [(f"\u0421\u0442\u0443\u0434\u0435\u043d\u04421.{i}", "") for i in range(1, 31)],
        rami: [(f"\u0421\u0442\u0443\u0434\u0435\u043d\u04422.{i}", "") for i in range(1, 31)],
        nemat: [(f"\u0421\u0442\u0443\u0434\u0435\u043d\u04423.{i}", "") for i in range(1, 31)],
    }
    all_students = [(practitioner, name, tg) for practitioner, group in group_rosters.items() for name, tg in group]
    roster, task_labels, task_values, result_values = extract_roster_and_sample(SOURCE)
    wb = Workbook()
    wb.remove(wb.active)

    common = wb.create_sheet("Общий")
    title(common, "Дискретная математика — общий лист", "Здесь хранится общий список студентов и итоговые результаты по трём практикам.", 13)
    common.append(["Практик", "Имя", "TG", "Активен", "D", "CW", "Тир", "Экзамен", "Оценка по таблице", "Финальная оценка", "Комментарий"])
    header_style(common, 3, 11)
    students = [("Артём", name, tg) for name, tg in roster]
    students += [("Артём", "", "")] * max(0, 30 - len(roster))
    students += [("Рами", "", "")] * 30
    students += [("Немат", "", "")] * 30
    students = all_students
    for row, (pr, name, tg) in enumerate(students, start=4):
        common.cell(row, 1, pr)
        common.cell(row, 2, name)
        common.cell(row, 3, tg)
        common.cell(row, 4, "Да")
        common.cell(row, 5, f'=IF(B{row}="","",SUMIF(\'Логи\'!$C$7:$C$3006,B{row},\'Логи\'!$H$7:$H$3006))')
        common.cell(row, 7, f'=IF(B{row}="","",IF(AND(E{row}>=$M$7,F{row}>=$M$8),"S",IF(AND(E{row}>=$M$5,F{row}>=$M$6),"A",IF(E{row}>=$M$4,"B","Допса"))))')
        for col in [4, 5, 6, 8, 9, 10, 11]:
            common.cell(row, col).protection = Protection(locked=False)
            common.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
    body_style(common, 4, 93, 1, 11)
    add_validation(common, "D4:D93", "Да,Нет", "Выберите статус студента: Да или Нет")
    add_validation(common, "H4:H93", "неуд,уд,хорошо,очень хорошо", "Выберите категорию экзамена")
    common["L3"] = "Порог"
    common["M3"] = "Значение"
    config = [
        ("Тир B: D", 17), ("Тир A: D", 35), ("Тир A: CW", 14),
        ("Тир S: D", 55), ("Тир S: CW", 24), ("Правило коэффициента 1", "> 2x/3"),
        ("Правило коэффициента 0.8", ">= x/2"), ("Коэффициент 1", 1),
        ("Коэффициент 0.8", 0.8), ("Коэффициент 0.5", 0.5),
    ]
    for row, (label, value) in enumerate(config, start=4):
        common.cell(row, 12, label)
        common.cell(row, 13, value)
    header_style(common, 3, 13)
    body_style(common, 4, 13, 12, 13)
    common.freeze_panes = "D4"
    common.auto_filter.ref = "A3:K93"
    common.sheet_view.showGridLines = False
    for col, width in {"A": 12, "B": 27, "C": 22, "D": 10, "E": 10, "F": 10, "G": 12, "H": 22, "I": 22, "J": 18, "K": 30, "L": 25, "M": 12}.items():
        common.column_dimensions[col].width = width
    common.protection.sheet = True
    common.protection.password = "common"

    cw = wb.create_sheet("CW")
    build_cw_sheet_template(cw, SOURCE, [(name, tg) for _practitioner, name, tg in all_students])
    ladder = wb.create_sheet("Логи")
    build_ladder_sheet(ladder)

    for practitioner_no, practitioner in enumerate(["Артём", "Рами", "Немат"], start=1):
        ws = wb.create_sheet(f"Плюсы_{practitioner}")
        build_plus_sheet(ws, practitioner, group_rosters[practitioner], task_labels, {}, {}, "Общий", practitioner_no)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUTPUT)
    print("Workbook created")


if __name__ == "__main__":
    build()
