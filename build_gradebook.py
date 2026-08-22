"""Small source workbook reader used by the compact workbook builder."""

from openpyxl import load_workbook


def extract_roster_and_sample(source_path):
    workbook = load_workbook(source_path, data_only=False)
    sheet = workbook.worksheets[0]

    roster = []
    for row in range(1, sheet.max_row + 1):
        name = sheet.cell(row, 1).value
        tg = sheet.cell(row, 2).value
        if name and str(name).strip().lower() != "имя":
            roster.append((str(name), "" if tg is None else str(tg)))

    task_labels = []
    for row in range(1, min(sheet.max_row, 10) + 1):
        if str(sheet.cell(row, 1).value).strip().lower() == "имя":
            for col in range(4, sheet.max_column + 1):
                value = sheet.cell(row, col).value
                if value is not None:
                    task_labels.append(str(value))
            break

    return roster, task_labels, {}, {}
