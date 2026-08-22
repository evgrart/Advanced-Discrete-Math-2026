from __future__ import annotations

import sys
from copy import copy
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / ".deps"))

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).parent
PRACTICES = 15
STUDENTS = 30
TASKS = 30
PRACTITIONERS = ["Артём", "Рами", "Немат"]

NAVY = "17365D"
PLUS_GREEN = "C6E0B4"
ONE_GREEN = "548235"
RED = "C00000"
WHITE = "FFFFFF"
INPUT = "FFFFFF"
THIN = Side(style="thin", color="D9E1F2")


def task_label(value):
    text = "" if value is None else str(value).strip()
    try:
        number = float(text.replace(",", "."))
        if number.is_integer():
            return str(int(number))
    except ValueError:
        pass
    return text


def cw_status_formula(row, first_row, last_row):
    parts = [
        f'COUNTIFS(\'CW\'!$A${first_row}:$A${last_row},A{row},\'CW\'!${col}${first_row}:${col}${last_row},">=3")'
        for col in "EFGHI"
    ]
    return f'=IF(A{row}="","",IF({"+".join(parts)}>=4,"зачет","незачет"))'


def final_grade_formula(row):
    """Return one final grade, resolving slash grades through the task result."""
    return (
        f'=IF(A{row}="","",IF(AND(G{row}="зачет",H{row}="зачет"),IF(I{row}="F","2F",IF(L{row}="","",'
        f'SWITCH(I{row}&"|"&L{row},'
        f'"B|неуд","2F","B|уд","3E","B|хорошо","3D",'
        f'"B|очень хорошо",IF(M{row}="решена","4C","3D"),'
        f'"A|неуд","2F","A|уд","3D","A|хорошо","4C",'
        f'"A|очень хорошо",IF(M{row}="решена","5A","4B"),'
        f'"S|неуд","2F","S|уд","3D","S|хорошо","4B",'
        f'"S|очень хорошо","5A",""))),"2F"))'
    )


def hide_top_rows(ws):
    ws.row_dimensions[1].hidden = True
    ws.row_dimensions[2].hidden = True


def header_style(ws, row, end_col):
    for col in range(1, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=WHITE))
    ws.row_dimensions[row].height = 28


def body_style(ws, first, last, first_col, last_col):
    for row in ws.iter_rows(min_row=first, max_row=last, min_col=first_col, max_col=last_col):
        for cell in row:
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def align_right(ws, first, last, columns):
    for col in columns:
        for row in range(first, last + 1):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)


def align_center(ws, first, last, columns):
    for col in columns:
        for row in range(first, last + 1):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_calibri(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            font = copy(cell.font)
            font.name = "Calibri"
            cell.font = font


def add_validation(ws, cell_range, values, message, kind="list"):
    if kind == "list":
        validation = DataValidation(type="list", formula1='"' + values + '"', allow_blank=True)
    else:
        validation = DataValidation(type=kind, operator="between", formula1="0", formula2="4", allow_blank=True)
    validation.errorStyle = "stop"
    validation.showErrorMessage = True
    validation.errorTitle = "Недопустимое значение"
    validation.error = message
    ws.add_data_validation(validation)
    validation.add(cell_range)


def color_tasks(ws, header_row, first_row, last_row):
    for col in range(6, 36):
        ws.cell(header_row, col).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(header_row, col).font = Font(bold=True, color="000000")
    ws.conditional_formatting.add(
        f"F{header_row}:AI{header_row}",
        FormulaRule(
            formula=[f'COUNTIF(F${first_row}:F${last_row},"1")>0'],
            fill=PatternFill("solid", fgColor=ONE_GREEN),
            font=Font(bold=True, color=WHITE),
        ),
    )
    rng = f"F{first_row}:AI{last_row}"
    anchor = f"F{first_row}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'{anchor}="+"'], fill=PatternFill("solid", fgColor=PLUS_GREEN), font=Font(color="000000")))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'{anchor}="1"'], fill=PatternFill("solid", fgColor=ONE_GREEN), font=Font(color=WHITE)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'{anchor}="-"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=WHITE)))


def make_common_legacy(wb):
    ws = wb.create_sheet("Общий")
    hide_top_rows(ws)
    headers = ["ФИО", "Практик", "TG", "Прод", "D", "CW", "КР-1", "КР-2", "Тир", "Бонус к CW", "Бонус к D", "Экзамен", "Задача", "Итоговая оценка"]
    for col, value in enumerate(headers, 1):
        ws.cell(3, col, value)
    header_style(ws, 3, 14)
    for row in range(4, 94):
        name = f"Студент{1 if row <= 33 else 2 if row <= 63 else 3}.{row - (3 if row <= 33 else 33 if row <= 63 else 63)}"
        practitioner = PRACTITIONERS[0 if row <= 33 else 1 if row <= 63 else 2]
        ws.cell(row, 1, name)
        ws.cell(row, 2, practitioner)
        ws.cell(row, 4, "Да")
        ws.cell(row, 5, f'=IF(A{row}="","",SUMIF(\'Логи\'!$A$4:$A$3006,A{row},\'Логи\'!$H$4:$H$3006))')
        ws.cell(row, 6, f'=IF(A{row}="","",SUMIF(\'CW\'!$A:$A,A{row},\'CW\'!$D:$D))')
        ws.cell(row, 7, cw_status_formula(row, 3, 92))
        ws.cell(row, 8, cw_status_formula(row, 96, 185))
        ws.cell(row, 9, f'=IF(A{row}="","",IF(AND(E{row}+IF(K{row}="",0,K{row})>=55,F{row}+IF(J{row}="",0,J{row})>=24),"S",IF(AND(E{row}+IF(K{row}="",0,K{row})>=35,F{row}+IF(J{row}="",0,J{row})>=14),"A",IF(E{row}+IF(K{row}="",0,K{row})>=17,"B","F"))))')
        ws.cell(row, 10, "")
        ws.cell(row, 11, "")
        ws.cell(row, 12).fill = PatternFill("solid", fgColor=INPUT)
        ws.cell(row, 11, f'=IF(OR(A{row}="",J{row}=""),"",IF(G{row}="B",IF(J{row}="неуд","2F",IF(J{row}="уд","3E",IF(J{row}="хорошо","3D","3D/4C"))),IF(G{row}="A",IF(J{row}="неуд","2F",IF(J{row}="уд","3D",IF(J{row}="хорошо","4C","4B/5A"))),IF(G{row}="S",IF(J{row}="неуд","2F",IF(J{row}="уд","3D",IF(J{row}="хорошо","4B/5A","5A"))),""))))')
        ws.cell(row, 13, "")
        ws.cell(row, 14, final_grade_formula(row))
        for col in [1, 2, 3, 4, 10, 11, 12, 13]:
            ws.cell(row, col).protection = Protection(locked=False)
            ws.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
    body_style(ws, 4, 93, 1, 14)
    align_right(ws, 4, 93, [1, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14])
    align_center(ws, 4, 93, [2, 9])
    ws.conditional_formatting.add("A4:N93", FormulaRule(formula=['AND($G4="зачет",$H4="зачет")'], fill=PatternFill("solid", fgColor=PLUS_GREEN), font=Font(color="000000")))
    add_validation(ws, "B4:B93", ",".join(PRACTITIONERS), "Выберите практика")
    add_validation(ws, "D4:D93", "Да,Нет", "Выберите Да или Нет")
    bonus = DataValidation(type="decimal", operator="between", formula1="-100", formula2="100", allow_blank=True)
    bonus.errorStyle = "stop"
    bonus.showErrorMessage = True
    bonus.error = "Введите числовой бонус"
    ws.add_data_validation(bonus)
    bonus.add("H4:I93")
    add_validation(ws, "J4:J93", "неуд,уд,хорошо,очень хорошо", "Выберите категорию экзамена")
    ws.freeze_panes = "D4"
    ws.auto_filter.ref = "A3:L93"
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 27, "B": 12, "C": 22, "D": 10, "E": 10, "F": 10, "G": 12, "H": 12, "I": 12, "J": 22, "K": 20, "L": 30}.items():
        ws.column_dimensions[col].width = width
    return ws


def make_common(wb):
    ws = wb.create_sheet("Общий")
    hide_top_rows(ws)
    headers = ["ФИО", "Практик", "TG", "Прод", "D", "CW", "КР-1", "КР-2", "Тир", "Бонус к CW", "Бонус к D", "Экзамен", "Задача", "Итоговая оценка"]
    for col, value in enumerate(headers, 1):
        ws.cell(3, col, value)
    header_style(ws, 3, 14)
    for row in range(4, 94):
        name = f"Студент{1 if row <= 33 else 2 if row <= 63 else 3}.{row - (3 if row <= 33 else 33 if row <= 63 else 63)}"
        practitioner = PRACTITIONERS[0 if row <= 33 else 1 if row <= 63 else 2]
        ws.cell(row, 1, name)
        ws.cell(row, 2, practitioner)
        ws.cell(row, 4, "Да")
        ws.cell(row, 5, f'=IF(A{row}="","",SUMIF(\'Логи\'!$A$4:$A$3006,A{row},\'Логи\'!$H$4:$H$3006))')
        ws.cell(row, 6, f'=IF(A{row}="","",SUMIF(\'CW\'!$A:$A,A{row},\'CW\'!$D:$D))')
        ws.cell(row, 7, cw_status_formula(row, 3, 92))
        ws.cell(row, 8, cw_status_formula(row, 96, 185))
        ws.cell(row, 9, f'=IF(A{row}="","",IF(AND(E{row}+IF(K{row}="",0,K{row})>=55,F{row}+IF(J{row}="",0,J{row})>=24),"S",IF(AND(E{row}+IF(K{row}="",0,K{row})>=35,F{row}+IF(J{row}="",0,J{row})>=14),"A",IF(E{row}+IF(K{row}="",0,K{row})>=17,"B","F"))))')
        ws.cell(row, 10, "")
        ws.cell(row, 11, "")
        ws.cell(row, 12).fill = PatternFill("solid", fgColor=INPUT)
        ws.cell(row, 13, "")
        ws.cell(row, 14, final_grade_formula(row))
        for col in [1, 2, 3, 4, 10, 11, 12, 13]:
            ws.cell(row, col).protection = Protection(locked=False)
            ws.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
    body_style(ws, 4, 93, 1, 14)
    align_right(ws, 4, 93, [1, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14])
    align_center(ws, 4, 93, [2, 9])
    ws.conditional_formatting.add("A4:N93", FormulaRule(formula=['AND($G4="зачет",$H4="зачет")'], fill=PatternFill("solid", fgColor=PLUS_GREEN), font=Font(color="000000")))
    add_validation(ws, "B4:B93", ",".join(PRACTITIONERS), "Выберите практика")
    add_validation(ws, "D4:D93", "Да,Нет", "Выберите Да или Нет")
    bonus = DataValidation(type="decimal", operator="between", formula1="-100", formula2="100", allow_blank=True)
    bonus.errorStyle = "stop"
    bonus.showErrorMessage = True
    bonus.error = "Введите числовой бонус"
    ws.add_data_validation(bonus)
    bonus.add("J4:K93")
    add_validation(ws, "L4:L93", "неуд,уд,хорошо,очень хорошо", "Выберите категорию экзамена")
    add_validation(ws, "M4:M93", "решена,не решена", "Выберите: решена или не решена")
    ws.freeze_panes = "D4"
    ws.auto_filter.ref = "A3:N93"
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 27, "B": 12, "C": 22, "D": 10, "E": 10, "F": 10, "G": 12, "H": 12, "I": 12, "J": 12, "K": 12, "L": 22, "M": 15, "N": 20}.items():
        ws.column_dimensions[col].width = width
    return ws


def make_cw(wb):
    ws = wb.create_sheet("CW")
    names = []
    for i, practitioner in enumerate(PRACTITIONERS):
        for j in range(1, STUDENTS + 1):
            names.append((f"Студент{i + 1}.{j}", practitioner))
    output = 1
    for kr in (1, 2):
        if kr == 1:
            title, header = 2, 1
            first = 3
        else:
            title, header = output, output + 1
            first = header + 1
        ws.cell(title, 1, f"КР-{kr}")
        ws.cell(title, 1).font = Font(size=14, bold=True, color=NAVY, name="Calibri")
        ws.cell(title, 1).fill = PatternFill("solid", fgColor="D9D9D9")
        ws.merge_cells(start_row=title, start_column=1, end_row=title, end_column=9)
        last = first + len(names) - 1
        for col, value in enumerate(["ФИО", "Практик", "Попытка", "S", "1", "2", "3", "4", "5"], 1):
            ws.cell(header, col, value)
        header_style(ws, header, 9)
        for row, (name, _practitioner) in enumerate(names, first):
            ws.cell(row, 1, name)
            ws.cell(row, 2, _practitioner)
            ws.cell(row, 3).fill = PatternFill("solid", fgColor="FFF2CC")
            ws.cell(row, 3).protection = Protection(locked=False)
            ws.cell(row, 4, f'=IF(A{row}="","",SUM(E{row}:I{row})*IF(C{row}="",1,1/2^(C{row}-1)))')
            for col in range(5, 10):
                ws.cell(row, col).protection = Protection(locked=False)
                ws.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
        body_style(ws, first, last, 1, 9)
        align_right(ws, first, last, [1, 3, 4, 5, 6, 7, 8, 9])
        align_center(ws, first, last, [2])
        ws.conditional_formatting.add(
            f"A{first}:A{last}",
            FormulaRule(formula=[f'COUNTIF(E{first}:I{first},">=3")>=4'], fill=PatternFill("solid", fgColor=PLUS_GREEN), font=Font(color="000000")),
        )
        add_validation(ws, f"C{first}:C{last}", "1,2,3,4", "Введите целое число от 1 до 4")
        ws.data_validations.dataValidation[-1].type = "whole"
        ws.data_validations.dataValidation[-1].operator = "between"
        ws.data_validations.dataValidation[-1].formula1 = "1"
        ws.data_validations.dataValidation[-1].formula2 = "4"
        score = DataValidation(type="decimal", operator="between", formula1="0", formula2="4", allow_blank=True)
        score.errorStyle = "stop"
        score.showErrorMessage = True
        score.error = "Введите число от 0 до 4"
        ws.add_data_validation(score)
        score.add(f"E{first}:I{last}")
        output = last + 2
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 27, "B": 14, "C": 12, "D": 12}.items():
        ws.column_dimensions[col].width = width
    for col in range(5, 10):
        ws.column_dimensions[get_column_letter(col)].width = 11
    return ws


def make_ranking(wb):
    ws = wb.create_sheet("Рейтинг")
    hide_top_rows(ws)
    for col, value in enumerate(["ФИО", "Практик", "D", "CW", "Тир"], 1):
        ws.cell(3, col, value)
    header_style(ws, 3, 6)
    body_style(ws, 4, 93, 1, 5)
    align_right(ws, 4, 93, [1, 3, 4])
    align_center(ws, 4, 93, [2, 5])
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 27, "B": 12, "C": 10, "D": 10, "E": 12}.items():
        ws.column_dimensions[col].width = width
    return ws


def make_logs(wb):
    ws = wb.create_sheet("Логи")
    hide_top_rows(ws)
    headers = ["ФИО", "Практик", "Практика", "Задача", "Номер выхода", "База D", "Коэффициент", "Начислено D", "Статус", "Поправка", "Дата"]
    for col, value in enumerate(headers, 1):
        ws.cell(3, col, value)
    header_style(ws, 3, 11)
    for row in range(4, 3007):
        ws.cell(row, 5, f'=IF(A{row}="","",COUNTIFS($A$4:A{row},A{row},$F$4:F{row},"<>",$I$4:I{row},"<>ушел на базу"))')
        ws.cell(row, 6, f'=IF(E{row}="","",IF(E{row}<=3,6,IF(E{row}<=6,5,IF(E{row}<=9,4,IF(E{row}<=12,3,IF(E{row}<=15,2,0))))))')
        ws.cell(row, 8, f'=IF(OR(F{row}="",G{row}=""),"",F{row}*G{row}+IF(J{row}="",0,J{row}))')
        for col in [1, 2, 3, 4, 7, 9, 10, 11]:
            ws.cell(row, col).protection = Protection(locked=False)
    body_style(ws, 4, 3006, 1, 11)
    align_right(ws, 4, 3006, [3, 4, 5, 6, 7, 8, 11])
    align_right(ws, 4, 3006, [1, 9, 10])
    align_center(ws, 4, 3006, [2])
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 27, "B": 12, "C": 12, "D": 14, "E": 14, "F": 10, "G": 14, "H": 14, "I": 28, "J": 12, "K": 18}.items():
        ws.column_dimensions[col].width = width
    return ws


def make_plus(path, practitioner, index):
    wb = Workbook()
    ws = wb.active
    ws.title = "Практика 1"
    roster = [(f"Студент{index}.{i}", "") for i in range(1, STUDENTS + 1)]
    labels = [str(i) for i in range(1, TASKS + 1)]
    for practice in range(1, PRACTICES + 1):
        if practice > 1:
            ws = wb.create_sheet(f"Практика {practice}")
        header = 3
        first, last = 4, 3 + STUDENTS
        hide_top_rows(ws)
        ws.cell(1, 1, f"Практика {practice} — плюсы")
        ws.cell(1, 1).font = Font(size=14, bold=True, color=NAVY)
        ws.cell(1, 1).fill = PatternFill("solid", fgColor="D9D9D9")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=35)
        for col, value in enumerate(["ФИО", "Присутствие", "S", "Коэффициент", "D за практику"] + labels, 1):
            ws.cell(header, col, value)
        header_style(ws, header, 35)
        for col, fill in [(1, "A9D18E"), (2, "D9EAD3"), (3, "F4B183"), (4, "FFF2CC"), (5, "D9EAD3")]:
            ws.cell(header, col).fill = PatternFill("solid", fgColor=fill)
            ws.cell(header, col).font = Font(bold=True, color="000000")
        add_validation(ws, f"B{first}:B{last}", "Да,Нет", "Выберите Да или Нет")
        for row, (name, tg) in enumerate(roster, first):
            ws.cell(row, 1, name)
            ws.cell(row, 3, f'=IF(A{row}="","",COUNTIF(F{row}:AI{row},"+")+COUNTIF(F{row}:AI{row},"1"))')
            ws.cell(row, 4, f'=IF(A{row}="","",IF(C{row}>2*COUNTA(F${header}:AI${header})/3,1,IF(C{row}>=COUNTA(F${header}:AI${header})/2,0.8,0.5)))')
            for col in range(6, 36):
                ws.cell(row, col).number_format = "@"
                ws.cell(row, col).protection = Protection(locked=False)
                ws.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
        body_style(ws, first, last, 1, 35)
        align_right(ws, first, last, [1, 3, 4, 5] + list(range(6, 36)))
        align_center(ws, first, last, [2])
        color_tasks(ws, header, first, last)
        ws.conditional_formatting.add(f"C{first}:C{last}", CellIsRule(operator="greaterThanOrEqual", formula=["21"], fill=PatternFill("solid", fgColor="70AD47")))
        ws.conditional_formatting.add(f"C{first}:C{last}", CellIsRule(operator="between", formula=["15", "20"], fill=PatternFill("solid", fgColor="A9D18E")))
        ws.conditional_formatting.add(f"C{first}:C{last}", CellIsRule(operator="lessThan", formula=["15"], fill=PatternFill("solid", fgColor="E2F0D9")))
        ws.freeze_panes = "F4"
        ws.sheet_view.showGridLines = False
        for col, width in {"A": 27, "B": 14, "C": 9, "D": 14, "E": 14}.items():
            ws.column_dimensions[col].width = width
        for col in range(6, 36):
            ws.column_dimensions[get_column_letter(col)].width = 10
        ws.protection.sheet = False
    make_plus_logs(wb)
    for sheet in wb.worksheets:
        set_calibri(sheet)
    wb.save(path)


def make_plus_logs(wb):
    ws = wb.create_sheet("Логи")
    headers = ["ФИО", "Практика", "Задача", "Действие", "Дата"]
    for col, value in enumerate(headers, 1):
        ws.cell(1, col, value)
    header_style(ws, 1, 5)
    for row in range(2, 3002):
        for col in range(1, 6):
            ws.cell(row, col).protection = Protection(locked=False)
    body_style(ws, 2, 3001, 1, 5)
    align_right(ws, 2, 3001, [1, 2, 3, 4, 5])
    ws.conditional_formatting.add("D2:D3001", FormulaRule(
        formula=['$D2="Поставил"'], fill=PatternFill("solid", fgColor="E2F0D9")))
    ws.conditional_formatting.add("D2:D3001", FormulaRule(
        formula=['$D2="Убрал"'], fill=PatternFill("solid", fgColor="F4CCCC")))
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 27, "B": 12, "C": 12, "D": 20, "E": 20}.items():
        ws.column_dimensions[col].width = width


def main():
    central = Workbook()
    central.remove(central.active)
    make_common(central)
    make_cw(central)
    make_ranking(central)
    make_logs(central)
    for sheet in central.worksheets:
        set_calibri(sheet)
    central.calculation.fullCalcOnLoad = True
    central.calculation.forceFullCalc = True
    central.calculation.calcMode = "auto"
    central.save(ROOT / "ДМобщее.xlsx")
    for index, practitioner in enumerate(PRACTITIONERS, 1):
        make_plus(ROOT / f"ДМ{practitioner}.xlsx", practitioner, index)
    print("Created four DM workbooks")


if __name__ == "__main__":
    main()
